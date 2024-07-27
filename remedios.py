import datetime
from openpyxl import Workbook, load_workbook # type: ignore

# Función para cargar los datos de los remedios desde un archivo Excel
def cargar_datos_excel():
    try:
        wb = load_workbook("lista_remedios.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            nombre, stock, dosis_diaria, fecha_reposicion = row
            remedios[nombre] = {'dosis_diaria': dosis_diaria, 'stock': stock}
    except FileNotFoundError:
        print("No se encontró el archivo Excel. Se creará uno nuevo.")

# Función para agregar remedios
def agregar_remedio():
    cargar_datos_excel()
    nombre = input("Ingrese el nombre del remedio: ")
    dosis_diaria = float(input("Ingrese la dosis diaria: "))
    stock_inicial = float(input("Ingrese la cantidad inicial en stock: "))
    
    if nombre in remedios:
        stock_actual = remedios[nombre]['stock']
        stock_total = stock_actual + stock_inicial
        remedios[nombre]['stock'] = stock_total
    else:
        remedios[nombre] = {'dosis_diaria': dosis_diaria, 'stock': stock_inicial}
    print("Remedio agregado exitosamente.")

# Función para generar archivo Excel con la lista de medicamentos
def generar_archivo_excel():
    wb = Workbook()
    ws = wb.active
    ws.append(["Nombre", "Stock", "Dosis diaria", "Fecha de reposición"])

    fecha_actual = datetime.date.today()
    for nombre, info in remedios.items():
        dias_para_reponer = (info['stock'] - 10) / info['dosis_diaria']
        fecha_reponer = fecha_actual + datetime.timedelta(days=dias_para_reponer)
        ws.append([nombre, info['stock'], info['dosis_diaria'], fecha_reponer.strftime('%d/%m/%Y')])

    nombre_archivo = f"lista_remedios.xlsx"
    wb.save(nombre_archivo)
    print(f"Archivo Excel generado: {nombre_archivo}")

# Función para mostrar la lista de todos los medicamentos ingresados
def mostrar_medicamentos():
    print("\nLista de medicamentos ingresados:")
    print("Nombre\t\tStock\tDosis diaria\tFecha de reposición")
    for nombre, info in remedios.items():
        dias_para_reponer = (info['stock'] - 10) / info['dosis_diaria']
        fecha_reponer = datetime.date.today() + datetime.timedelta(days=dias_para_reponer)
        print(f"{nombre}\t\t{info['stock']}\t{info['dosis_diaria']}\t\t{fecha_reponer.strftime('%d/%m/%Y')}")

# Función para el menú de la aplicación
def menu():
    while True:
        print("\nMenú:")
        print("1. Agregar remedio")
        print("2. Generar archivo Excel con la lista de medicamentos")
        print("3. Mostrar lista de medicamentos")
        print("4. Salir")
        opcion = input("Seleccione una opción: ")

        if opcion == "1":
            agregar_remedio()
        elif opcion == "2":
            generar_archivo_excel()
        elif opcion == "3":
            mostrar_medicamentos()
        elif opcion == "4":
            print("Saliendo...")
            break
        else:
            print("Opción no válida. Por favor, seleccione una opción válida.")

# Ejemplo de uso
if __name__ == "__main__":
    remedios = {}
    menu()
